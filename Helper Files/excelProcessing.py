

# Basic Modules
import os
import sys
import numpy as np
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font



class dataProcessing:        
        
    def convertToXLSX(self, pulseExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(pulseExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return pulseExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(pulseExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(pulseExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = pulseExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile) as newFile:
                reader = csv.reader(newFile, delimiter = excelDelimiter)
                for row in reader:
                    xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the GSR Data from the Excel File
        else:
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    

class processPulseData(dataProcessing):
    
    def getData(self, pulseExcelFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            pulseExcelFile: The Path to the Excel File Containing the Pulse Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(pulseExcelFile):
            print("The following Input File Does Not Exist:", pulseExcelFile)
            sys.exit()
        # Convert to Exel if .xls Format; If .xlsx, Do Nothing; If Other, Exit Program
        pulseExcelFile = self.convertToXLSX(pulseExcelFile)

        print("Extracting Data from the Excel File:", pulseExcelFile)
        # Load Data from the Excel File
        WB = xl.load_workbook(pulseExcelFile, data_only=True, read_only=True)
        WB_worksheets = WB.worksheets
        ExcelSheet = WB_worksheets[testSheetNum]
        
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) == type(1.1):
                dataStartRow = cellA.row
                break
        
        data = dict(time=[], Capacitance=[])
        # Loop Through the Excel Worksheet to collect all the data
        for pointNum, [colA, colB] in enumerate(ExcelSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=2, max_row=ExcelSheet.max_row)):
            # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
            Time = colA.value; Capacitance = colB.value;
            
            # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
            if Time == None or Capacitance == None:
                break
            
            # Add Data to Dictionary
            data["time"].append(Time)
            data["Capacitance"].append(Capacitance)
             
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Data Collecting"); WB.close()
        return np.array(data["time"]), np.array(data["Capacitance"])


    def saveResults(self, bloodPulse, pulseNumSaving, saveDataFolder, saveExcelName, sheetName = "Blood Pulse Data"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excel_file = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excel_file):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excel_file)
            WB_worksheet = WB.create_sheet(sheetName)
        
        # Label First Row
        header = ["Pulse Number", "Start Time", "End Time", 'Systolic Time From Start', 'Tidal Wave Time From Systolic', 'Dicrotic Time From Tidal Wave',
                  'Tail Wave Time From Dicrotic', 'End Time From Tail Wave', 'Systolic Peak Amplitude', 'Tidal Wave Peak Ampltiude',
                  "Dicrotic Peak Amplitude", 'Tail Wave Peak Ampltitude']
        header.extend(["", 'Gaussian Systolic Time From Start', 'Gaussian Tidal Wave Time From Systolic', 'Gaussian Dicrotic Time From Tidal Wave',
                  'Gaussian Tail Wave Time From Dicrotic', 'Gaussian End Time From Tail Wave', 'Gaussian Systolic Peak Amplitude', 'Gaussian Tidal Wave Peak Ampltiude',
                  "Gaussian Dicrotic Peak Amplitude", 'Gaussian Tail Wave Peak Ampltitude'])
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for pulseNum in pulseNumSaving:
            # Write the Data to Excel
            WB_worksheet.append(bloodPulse[pulseNum]["Results to Save"])
        
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
            
        # Save as New Excel File
        WB.save(excel_file)
        WB.close()
    
    
class processGSRData(dataProcessing):
    
    def extractCHIData_CurrentTime(self, chiWorksheet):
        
        # -------------------------------------------------------------------#
        # ----------------------- Extract Run Info --------------------------#
        
        # Get Time and Current Data from Excel as Well as CHI Labeled Peaks
        findStart = True
        timePoints = []; currentPoints = []
        peakTimesCHI = []; peakCurrentsCHI = []; peakAmplitudesCHI = []
        # Loop Through the Info Section and Extract the Needxed Run Info from Excel
        rowGenerator = chiWorksheet.rows
        for cell in rowGenerator:
            # Get Cell Value
            cellVal = cell[0].value
            
            # Extract Improtant Information from the File
            if findStart:
                # If Nothing in Cell, Continue
                if cellVal == None:
                    continue
                # If Time Peak Found by CHI, Store the Value
                elif cellVal.startswith("tp = "):
                    peakTimeVal = float(cellVal.split(" = ")[-1][:-1])
                    peakTimesCHI.append(peakTimeVal)
                # If Time Peak Found by CHI, Store the Value
                elif cellVal.startswith("ip = "):
                    peakCurrentVal = float(cellVal.split(" = ")[-1][:-1])
                    peakCurrentsCHI.append(peakCurrentVal)
                # If Amplitude Peak Found by CHI, Store the Value
                elif cellVal.startswith("Ap = "):
                    peakAmplitudeVal = float(cellVal.split(" = ")[-1][:-1])
                    peakAmplitudesCHI.append(peakAmplitudeVal)
                # If Current/Time Titles are Present, the Data is Starting Soon
                elif cellVal == "Time/sec":
                    next(rowGenerator) # Skip Over Empty Cell After Title
                    findStart = False
            # Extract the Data from the File
            else:
                # Break out of Loop if no More Data (edge effect if someone edits excel)
                if cell[0].value == None:
                    break
                
                # Keep Track  the Time and Current Data points
                timePoints.append(float(cell[0].value))
                currentPoints.append(float(cell[1].value))
        
        # Return Time and Current
        currentPoints = np.array(currentPoints)
        timePoints = np.array(timePoints)
        return timePoints, currentPoints

    def getData(self, gsrFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            gsrFile: The Path to the Excel/TXT/CSV File Containing the GSR Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(gsrFile):
            print("The following Input File Does Not Exist:", gsrFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if gsrFile.endswith(".txt") or gsrFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(gsrFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(gsrFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, chiWorksheet = self.convertToExcel(gsrFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif gsrFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(gsrFile, data_only=True, read_only=True)
            chiWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", gsrFile)
        print("Extracting Data from the Excel File:", gsrFile)
        
        # Extract Time and Current Data from the File
        timePoints, currentPoints = self.extractCHIData_CurrentTime(chiWorksheet)
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting GSR Data");
        return timePoints, currentPoints
    
    def saveResults(self, timeGSR, currentGS, saveDataFolder, saveExcelName, sheetName = "Galvanic Skin Response Data"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excel_file = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excel_file):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excel_file)
            WB_worksheet = WB.create_sheet(sheetName)
        
        # Label First Row
        header = ["Time (Seconds)", "Current (uAmps)"]
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for dataPoint in range(len(timeGSR)):
            timePoint = timeGSR[dataPoint]
            currentPoint = currentGS[dataPoint]
            
            # Write the Data to Excel
            WB_worksheet.append([timePoint, currentPoint])
        
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
            
        # Save as New Excel File
        WB.save(excel_file)
        WB.close()

        
  
    
    